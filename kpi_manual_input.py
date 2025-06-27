from flask_restful import Resource, reqparse
from flask import jsonify, request, send_file, make_response
from sqlalchemy import or_, and_, insert, select, func, update, cast, DateTime
from sqlalchemy.orm import aliased
from sqlalchemy import and_, select, func, Integer, cast
from db import db, Base, metadata
from config import db_schema
import ast
from authentication import (
    user_or_admin_authentication_required,
    admin_authentication_required,
    get_user_from_token,
)
import datetime
from datetime import datetime, timedelta,date
import pandas as pd

from sqlalchemy.exc import IntegrityError
import math
from math import isnan
from resources.kpi_management import FetchSectorRole

from utils import max_day_and_days_difference
from schemas.kpi_hierarchy_schema import (
    hierarchy_schemas,
    org_level_schemas,
)
from schemas.kpi_search_schema import search_schemas
from schemas.kpi_input_schema import kpi_input_schema

# from schemas.kpi_hier_mapping_schema import org_hier_mapping_schema, org_hier_mapping_schemas
from resources.resources_utils.kpi_manual_input_util import (
    fetch_kpi_input_histroy,
    insert_kpi_input_history,
)
from helper.kpi_manual_input_helper import kpi_manual_input
from resources.kpi_management import KPINonRollup, filter_view_based_data
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from helper.org_process_mapping_helper import OrgProcessMappingHelper
from helper.org_hierarchy_mapping_helper import OrgHierarchyMappingHelper
dateformatStr = "%Y-%m-%d"


class OrganizationKPI(Resource):
    def __init__(self):
        super().__init__()
        self.default_date = '9999-01-01 00:00:00.000'

    @user_or_admin_authentication_required
    def get(self):
        """
        GET API for listing organization hierarchy.
        :param: org_id, hier_type, IsDeactivation - not mandatory
        :return: list of hierarchy data
        """
        org_hierarchy = Base.metadata.tables['db_nxtgen.Org_Hierarchy']
        q = request.args.get('org_id')
        hier_type = request.args.get('hier_type')
        is_deactivation_param = request.args.get("IsDeactivation", "").strip().lower()
        if is_deactivation_param in ["true", "1", "yes"]:
            is_deactivation = True
        elif is_deactivation_param in ["false", "0", "no"]:
            is_deactivation = False
        else:
            is_deactivation = None  # No deactivation filter provided

        # Initializing the base query
        query = db.session.query(org_hierarchy)

        # Adding filters based on parameters
        if q:
            org_ids = ast.literal_eval(q)
            if not isinstance(org_ids, list):
                org_ids = [org_ids]
            query = query.filter(org_hierarchy.c.org_id.in_(org_ids))
        if hier_type:
            query = query.filter(org_hierarchy.c.hier_type == hier_type)
        else:
            query = query.filter(org_hierarchy.c.hier_type == 'Standard')

        if is_deactivation is None:
           

            # Activated records query
            activated_query = query.filter(
                and_(
                    org_hierarchy.c.record_cut_over_date == self.default_date,
                    org_hierarchy.c.rec_end_date == self.default_date
                )
            )

            # Deactivated records  
            deactivated_query = query.filter(
                and_(
                    cast(org_hierarchy.c.record_cut_over_date, DateTime) < func.current_date(),
                    org_hierarchy.c.rec_end_date == self.default_date
                )
            )

         
            activated_records = self.deactivation_activation_records(activated_query, False)
            deactivated_records = self.deactivation_activation_records(deactivated_query, True)

           
            all_records = activated_records + deactivated_records

            return all_records, 200
        elif not is_deactivation:
            # Activated records (IsDeactivation is false)
            query = query.filter(
                and_(
                    org_hierarchy.c.record_cut_over_date == self.default_date,
                    org_hierarchy.c.rec_end_date == self.default_date
                )
            )
        else:
            # Deactivated records (IsDeactivation is true)
            query = query.filter(
                and_(
                    cast(org_hierarchy.c.record_cut_over_date, DateTime) < func.current_date(),
                    org_hierarchy.c.rec_end_date == self.default_date
                )
            )

       
        return self.deactivation_activation_records(query, is_deactivation), 200

    def deactivation_activation_records(self, query, type_records):
        """
        Helper method to handle deactivation/activation records.
        """
        records = db.session.execute(query).fetchall()

        serialized_records = hierarchy_schemas.dump(records)

        deactivation_activation_orgs = [
            {
                "is_deactivation": type_records,
                **record
            }
            for record in serialized_records
        ]
        return deactivation_activation_orgs



class OrganizationManagement(Resource):

    def __init__(self):
        # Initialize constants and table references for DB interactions
        self.DEFAULT_REC_END_DATE = "9999-01-01 00:00:00.000"
        self.org_hierarchy_table = Base.metadata.tables["db_nxtgen.Org_Hierarchy"]
        self.workflow_table = Base.metadata.tables["db_nxtgen.Workflow"]
        self.workflow_static_table = Base.metadata.tables["db_nxtgen.Workflow_Static"]
        self.org_hier_mapping_table = Base.metadata.tables["db_nxtgen.Org_Hier_Mapping"]
        self.process_area_mapping_table = Base.metadata.tables[
            "db_nxtgen.Process_Area_Mapping"
        ]
        self.org_process_mapping_helper = OrgProcessMappingHelper()
        self.org_hier_mapping_helper = OrgHierarchyMappingHelper()

    @user_or_admin_authentication_required
    def post(self):
        
        
        """
        Handles POST requests to create org hierarchy & workflow entries.
        Expects JSON input to manage organizations and their relationships.
        """
        try:
            data = request.get_json()
            print('request data....',data)
            response_messages = []

            for entry in data:
                # Process creation of new orgs
                if entry["action"] == "create":
                    if entry["prop_type"] == "org_name" and entry["type"] == "single":
                    
                        # Get user ID from token for tracking changes
                        print('get_user_from_token',get_user_from_token())
                        user_id = get_user_from_token()[1]
                        email_id = get_user_from_token()[3]
                        # Generate a new workflow ID based on existing max
                        max_wid = db.session.execute(
                            select(func.max(func.cast(func.substring(self.workflow_table.c.wid, 4, 1000), Integer)))
                            .where(self.workflow_table.c.wid.like("wf_%"))
                        ).scalar()
                        new_wid = f"wf_{(max_wid + 1) if max_wid else 1}"

                        for org_data in entry["prop_val"]:
                            # Determine the next org_id based on existing entries
                            mx_org_id_result = db.session.execute(
                                select(func.max(self.org_hierarchy_table.c.org_id))
                            ).scalar()
                            mx_org_id = (
                                (mx_org_id_result + 1) if mx_org_id_result else 1
                            )

                            level_columns = {
                                "org_id": mx_org_id,
                                "org_name": org_data["org_name"],
                                "h_id": f"{mx_org_id + 1000:04}",
                                "hier_type": org_data["hier_type"],
                                "org_level": org_data["org_level"],
                                "created_by": user_id,
                                "created_date": db.func.now(),
                                "rec_start_date": db.func.now(),
                                "rec_end_date": None,
                                "updated_date": None,
                                "approval_1_status": "Pending",
                                "approval_2_status": "Pending",
                                "is_deleted": 0
                               
                            }

                            # If a parent organization is specified, validate and fetch its details
                            if org_data.get("parent_name"):
                                parent_org_record = db.session.execute(
                                    select(self.org_hierarchy_table).where(
                                        self.org_hierarchy_table.c.org_name
                                        == org_data["parent_name"]
                                    )
                                ).fetchone()
                                if parent_org_record:
                                    parent_org_record = dict(parent_org_record._mapping)
                                    if (
                                        int(parent_org_record["org_level"])
                                        != level_columns["org_level"] - 1
                                    ):
                                        response_messages.append(
                                            f"Parent organization level is invalid for '{level_columns['org_name']}'."
                                        )
                                        continue

                                    # Dynamically populate levels based on parent org
                                    for level in range(
                                        1, int(level_columns["org_level"])
                                    ):
                                        column_name = f"level{level}"
                                        level_columns[column_name] = (
                                            parent_org_record.get(column_name)
                                        )

                                level_columns[
                                    f'level{level_columns["org_level"]}'
                                ] = org_data["org_name"]
                            else:
                                # Set level1 as org_name and org_level as 1 if no parent
                                level_columns["org_level"] = 1
                                level_columns["level1"] = level_columns["org_name"]

                            # Insert the new org record and get the record_id
                            result = db.session.execute(
                                self.org_hierarchy_table.insert()
                                .values(level_columns)
                                .returning(self.org_hierarchy_table.c.record_id)
                            )
                            record_id = result.scalar()
                            # Fetch mapping from Workflow_Static for creating a workflow entry
                            workflow_entry = db.session.execute(
                                select(self.workflow_static_table).where(
                                    self.workflow_static_table.c.action == "create",
                                    self.workflow_static_table.c.type == "single",
                                    self.workflow_static_table.c.prop_type
                                    == "organization",
                                    self.workflow_static_table.c.is_new == 1,
                                )
                            ).fetchone()

                            if workflow_entry:
                                
                                workflow_data = {
                                    "wid": new_wid,
                                    "prop_id": workflow_entry.prop_id,
                                    "org_hierarchy_record_id": record_id,
                                    "dyn_col_1": "org_name",
                                    "dyn_col_2": org_data["org_name"],
                                    "dyn_col_3": None,
                                    "dyn_col_4": None,
                                    "created_date": db.func.now(),
                                    "updated_date": db.func.now(),
                                    "is_deleted": 0,
                                    "wf_status": "Pending",
                                    "wf_reviewer_1_name": None,
                                    "wf_reviewer_2_name": None,
                                    "requested_by": user_id,
                                    "wf_reviewer_1_status": "Pending",
                                    "wf_reviewer_1_status_date": None,
                                    "wf_reviewer_2_status": "Pending",
                                    "wf_reviewer_2_status_date": None,
                                    "typeof_action": "create",
                                    "typeof_cr": "single",
                                    "Comments": None,
                                    "email_id":email_id
                                }
                                db.session.execute(
                                    self.workflow_table.insert().values(
                                        workflow_data
                                    )
                                )

                            response_messages.append(
                                f"{level_columns['org_name']} has been successfully added at level {level_columns['org_level']}"
                            )
                    

                    elif entry["prop_type"] == "hierarchy":
                        # Handle hierarchy mapping for given entries
                        max_wid = ""
                        for hierarchy_data in entry.get("prop_val", []):

                            if (
                                "source_system_cd" not in hierarchy_data
                                or "process_area" not in hierarchy_data
                            ):
                                response_messages.append(
                                    "source_system_cd and process_area are required."
                                )
                                continue

                            # Generate new workflow ID for hierarchy mappings
                            
                            result = db.session.execute(
                                select(self.workflow_table.c.wid.distinct()).where(
                                    self.workflow_table.c.org_hierarchy_record_id.in_(hierarchy_data.get("record_id"))
                                )
                            ).fetchall()
                            find_wid = [row[0] for row in result]

                            

                            max_wid = db.session.execute(
                                select(func.max(func.cast(func.substring(self.workflow_table.c.wid, 4, 1000), Integer)))
                                .where(self.workflow_table.c.wid.like("wf_%"))
                            ).scalar()
 
                            print("i am wid", max_wid )
                            new_wid = f"wf_{(max_wid + 1) if max_wid else 1}"

                            # Determine the last non-null level from hierarchy data
                            # last_non_null_level = None
                            last_non_null_level = None

                            for i in range(1, 8):
                                level_key = f"level{i}"
                                
                                # Check if the key exists in hierarchy_data and has a valid value
                                # Determine the last non-null level from hierarchy data
                            last_non_null_level = None
                            for i in range(1, 8):
                                level_key = f'level{i}'
                                if level_key in hierarchy_data and hierarchy_data[level_key]:
                                    last_non_null_level = hierarchy_data[level_key]

                            # Debugging output for the last non-null level
                            print('last_non_null_level:', last_non_null_level)

                            # Handle the case where no valid levels are found
                            if last_non_null_level is None:
                                response_messages.append("No valid levels found.")
                                continue

                            print("Org", last_non_null_level)
                                
                            org_record = db.session.execute(
                                select(self.org_hierarchy_table).where(
                                    self.org_hierarchy_table.c.org_name
                                    == last_non_null_level
                                )
                            ).fetchone()
                            if not org_record:
                                response_messages.append(
                                    f"No organization found with name '{last_non_null_level}'."
                                )
                                continue

                            org_id = org_record.org_id
                            is_exists_hierarchy = (
                                db.session.query(self.org_hier_mapping_table)
                                .filter_by(org_id=org_id,source_system_cd=hierarchy_data['source_system_cd'])
                                .first()
                            )
                            print('is_exists_hierarchy.....',is_exists_hierarchy)
                            # if is_exists_hierarchy:
                            #     return {
                            #         "message": "mapping for this organization already exists"
                            #     }, 400
                            # Fetch mapping ID based on source_system_cd and process_area
                            mapping_record = db.session.execute(
                                select(self.process_area_mapping_table).where(
                                    self.process_area_mapping_table.c.source_system_cd
                                    == hierarchy_data["source_system_cd"],
                                    self.process_area_mapping_table.c.process_area
                                    == hierarchy_data["process_area"],
                                )
                            ).fetchone()

                            if not mapping_record:
                                response_messages.append(
                                    f"No mapping found for source system '{hierarchy_data['source_system_cd']}' and process area '{hierarchy_data['process_area']}'."
                                )
                                continue

                            mapping_record = dict(mapping_record._mapping)
                            mapping_id = mapping_record["mapping_id"]
                            mx_id_result = db.session.execute(
                                    select(func.max(self.org_hier_mapping_table.c.id))
                                ).scalar()
                            # Prepare dynamic columns based on fetched mapping
                            dynamic_columns = {
                                "mapping_id": mapping_id,
                                "id": mx_id_result + 1 if mx_id_result !=None else 1,
                                "org_id": org_id,
                                "source_system_cd": hierarchy_data["source_system_cd"],
                                "created_date": db.func.now(),
                                "created_by": get_user_from_token()[1],
                                "updated_date": None,
                                "updated_by": None,
                                "is_deleted": 0,
                                "rec_start_date": db.func.now(),
                                "rec_end_date": None,
                                "approval_1_by": None,
                                "approval_1_date": None,
                                "approval_1_status": "Pending",
                                "approval_2_by": None,
                                "approval_2_date": None,
                                "approval_2_status": "Pending",
                            }

                            # Map values from hierarchy data to dynamic columns
                            for i in range(1, 13):
                                dynamic_field_name = f"dynamic_mapping_field_name_{i}"
                                if dynamic_field_name in mapping_record:
                                    key = mapping_record[dynamic_field_name]
                                    if key in hierarchy_data.get("mapping", {}):
                                        dynamic_columns[dynamic_field_name] = (
                                            hierarchy_data["mapping"][key]
                                        )

                            # Insert into Org_Hier_Mapping and get record_id
                            result = db.session.execute(
                                self.org_hier_mapping_table.insert()
                                .values(dynamic_columns)
                                .returning(self.org_hier_mapping_table.c.record_id)
                            )
                            record_id = result.scalar()
                            # Prepare entries for the Workflow table based on hierarchy data
                            workflow_data_entries = []
                            for field_name, value in hierarchy_data.get(
                                "mapping", {}
                            ).items():
                                workflow_data_entries.append(
                                    {
                                        "wid": new_wid,
                                        "mapping_record_id": record_id,
                                        "dyn_col_1": field_name,
                                        "dyn_col_2": value,
                                        "created_date": db.func.now(),
                                        "email_id": get_user_from_token()[3],
                                        "is_deleted": 0,
                                        "wf_status": "Pending",
                                        "wf_reviewer_1_status": "Pending",
                                        "wf_reviewer_2_status": "Pending",
                                        "requested_by": dynamic_columns["created_by"],
                                        "typeof_action": "create",
                                        "typeof_cr": "single",
                                        "Comments": None,
                                        "depedent_workflow_id": ",".join(find_wid)
                                    }
                                )

                            # Add process_area and source_system_cd entries
                            workflow_data_entries.append(
                                {
                                    "wid": new_wid,
                                    "mapping_record_id": record_id,
                                    "dyn_col_1": "process_area",
                                    "dyn_col_2": hierarchy_data["process_area"],
                                    "created_date": db.func.now(),
                                    "email_id": get_user_from_token()[3],
                                    "is_deleted": 0,
                                    "wf_status": "Pending",
                                    "wf_reviewer_1_status": "Pending",
                                    "wf_reviewer_2_status": "Pending",
                                    "requested_by": dynamic_columns["created_by"],
                                    "typeof_action": "create",
                                    "typeof_cr": "single",
                                    "Comments": None,
                                    "depedent_workflow_id": ",".join(find_wid)
                                }
                            )
                            workflow_data_entries.append(
                                {
                                    "wid": new_wid,
                                    "mapping_record_id": record_id,
                                    "dyn_col_1": "source_system_cd",
                                    "dyn_col_2": hierarchy_data["source_system_cd"],
                                    "created_date": db.func.now(),
                                    "email_id": get_user_from_token()[3],
                                    "is_deleted": 0,
                                    "wf_status": "Pending",
                                    "wf_reviewer_1_status": "Pending",
                                    "wf_reviewer_2_status": "Pending",
                                    "requested_by": dynamic_columns["created_by"],
                                    "typeof_action": "create",
                                    "typeof_cr": "single",
                                    "Comments": None,
                                    "depedent_workflow_id": ",".join(find_wid)
                                }
                            )

                            # Include levels from level1 to level7 in workflow entries
                            for i in range(1, 8):
                                level_key = f"level{i}"
                                if level_key in hierarchy_data and hierarchy_data[level_key]:
                                    workflow_data_entries.append(
                                        {
                                            "wid": new_wid,
                                            "mapping_record_id": record_id,
                                            "dyn_col_1": level_key,
                                            "dyn_col_2": hierarchy_data.get(level_key),
                                            "created_date": db.func.now(),
                                            "email_id": get_user_from_token()[3],
                                            "is_deleted": 0,
                                            "wf_status": "Pending",
                                            "wf_reviewer_1_status": "Pending",
                                            "wf_reviewer_2_status": "Pending",
                                            "requested_by": dynamic_columns["created_by"],
                                            "typeof_action": "create",
                                            "typeof_cr": "single",
                                            "Comments": None,
                                            "depedent_workflow_id": ",".join(find_wid)
                                        }
                                )
                                    
                            # Insert all prepared workflow entries into the Workflow table
                            for workflow_data in workflow_data_entries:
                                db.session.execute(
                                    self.workflow_table.insert().values(workflow_data)
                                )

                            response_messages.append(
                                f"Hierarchy mapping for '{hierarchy_data['source_system_cd']}' has been successfully added with record ID {record_id}."
                            )

                db.session.commit()

                return {"messages": response_messages}, 200

        except Exception as e:
            # Rollback to clean up the session
            db.session.rollback()
            return {
                "message": "Something went wrong",
                "data": None,
                "error": str(e),
            }, 500

    @user_or_admin_authentication_required
    def put(self):

        try:
            current_user = get_user_from_token()[1]
            email_id = get_user_from_token()[3]
            data = request.get_json()
            
            if data["typeof_action"] == "name_change":

                max_wid = db.session.execute(
                    select(
                        func.max(
                            cast(
                                func.substring(self.workflow_table.c.wid, 4, 7),
                                Integer,
                            )
                        )
                    ).where(self.workflow_table.c.wid.like("wf_%"))
                ).scalar()
                new_wid = f"wf_{(int(max_wid) + 1) if max_wid else 1}"

                accumulated_message = []
                final_message = ""
                for sindata in data["data"]:
                    try:

                        result = db.session.query(
                            self.org_hierarchy_table.c.org_name, 
                            self.org_hierarchy_table.c.org_level
                        ).filter(
                            self.org_hierarchy_table.c.org_id == sindata["Org_id"],
                            # self.org_hierarchy_table.c.rec_end_date == '9999-01-01 00:00:00.000',
                            self.org_hierarchy_table.c.is_deleted == False
                        ).all()
                        if result:
                            org_name, org_level = result[0]

                        
                            level_column = f'level{org_level}'
                            
                            
                            query_count = db.session.query(self.org_hierarchy_table).filter(
                                self.org_hierarchy_table.c.org_id != sindata["Org_id"],
                                getattr(self.org_hierarchy_table.c, level_column) == org_name
                            ).count()
                            # Perform validation for organization name existence
                            is_org_name_exists = self.org_process_mapping_helper.validation_organization_match_score(
                                sindata["action_val"]
                            )
                            if is_org_name_exists:
                                return is_org_name_exists, 400

                            # Fetch the existing organization record
                            existing_record = db.session.execute(
                                select(self.org_hierarchy_table).where(
                                    self.org_hierarchy_table.c.org_id == sindata["Org_id"],
                                    self.org_hierarchy_table.c.is_deleted == 0,
                                    # self.org_hierarchy_table.c.rec_end_date == "9999-01-01 00:00:00.000",
                                )
                            ).fetchone()

                            if existing_record is None:
                                return {"message": "Organization not found."}, 404

                            # Convert fetched record to a dictionary
                            existing_record = dict(existing_record._mapping)
                            current_org_name = existing_record["org_name"]
                            org_level = existing_record["org_level"]
                            level_column_key = f"level{org_level}"

                            # Prepare new record data
                            new_record = {
                                "org_id": sindata["Org_id"],
                                "org_name": sindata["action_val"],
                                "hier_type": existing_record["hier_type"],
                                "org_level": org_level,
                                "level1": existing_record["level1"],
                                "level2": existing_record["level2"],
                                "level3": existing_record["level3"],
                                "level4": existing_record["level4"],
                                "level5": existing_record["level5"],
                                "level6": existing_record["level6"],
                                "level7": existing_record["level7"],
                                **{level_column_key: sindata["action_val"]},
                                "created_date": datetime.now(),
                                "created_by": current_user,
                                "updated_date": datetime.now(),
                                "updated_by": current_user,
                                "is_deleted": 0,
                                "rec_start_date": existing_record["rec_start_date"],
                                "rec_end_date": None,
                                "approval_1_by": None,
                                "approval_1_status": "Pending",
                                "approval_1_date": None,
                                "approval_2_by": None,
                                "approval_2_status": "Pending",
                                "approval_2_date": None,
                                "is_leaf_node": False if query_count > 0 else True,
                                "snapshot_refresh":sindata["snapshot_refresh"],
                                "snapshot_refresh_date": (
                                    datetime.strptime(sindata["snapshot_refresh_date"], "%Y-%m-%dT%H:%M:%S.%fZ")
                                    .strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if sindata.get("snapshot_refresh_date") else None
                                )
                            }

                            # Insert new record into the organization hierarchy table
                            result = db.session.execute(
                                self.org_hierarchy_table.insert()
                                .values(new_record)
                                .returning(self.org_hierarchy_table.c.record_id)
                            )
                            record_id = result.scalar()
                            db.session.commit()

                            # query to update the is_parent_pending column based on org_level and level column        
                            level_column = f"level{org_level}"         
                            db.session.execute(            
                                self.org_hierarchy_table.update()            
                                .where(                
                                    getattr(self.org_hierarchy_table.c, level_column) == current_org_name, 
                                    self.org_hierarchy_table.c.org_level >= org_level, 
                                    # self.org_hierarchy_table.c.org_id != sindata["Org_id"], 
                                    self.org_hierarchy_table.c.is_deleted == 0 
                                    ) 
                                .values(is_parent_pending='True') 
                                ) 
                            db.session.commit()

                            # Get workflow entry for update action
                            workflow_entry = db.session.execute(
                                select(self.workflow_static_table).where(
                                    self.workflow_static_table.c.action == "update",
                                    self.workflow_static_table.c.type == "single",
                                    self.workflow_static_table.c.prop_type == "organization",
                                    self.workflow_static_table.c.is_new == 0,
                                )
                            ).fetchone()

                            # Prepare workflow entry data
                            work_flow_entries = {
                                "wid": new_wid,  # Use the same workflow ID for all requests
                                "dyn_col_1": "org_name",
                                "dyn_col_2": current_org_name,
                                "dyn_col_3": sindata["action_val"],
                                "created_date": db.func.now(),
                                "email_id":email_id,
                                "updated_date": None,
                                "is_deleted": 0,
                                "wf_status": "Pending",
                                "wf_reviewer_1_name": None,
                                "wf_reviewer_2_name": None,
                                "requested_by": current_user,
                                "wf_reviewer_1_status": "Pending",
                                "wf_reviewer_1_status_date": None,
                                "wf_reviewer_2_status": "Pending",
                                "wf_reviewer_2_status_date": None,
                                "typeof_action": "update",
                                "typeof_cr": "single",
                                "Comments": None,
                                "depedent_workflow_id": None,
                                "prop_id": None,
                                "org_hierarchy_record_id": record_id,
                                "mapping_record_id": None,
                                "prop_id": workflow_entry.prop_id,
                            }

                            # Insert workflow entry
                            db.session.execute(
                                self.workflow_table.insert().values(work_flow_entries)
                            )
                            db.session.commit()
                            accumulated_message.append( f"{existing_record['org_name']} has been successfully updated to {sindata['action_val']}")
                    
                    except Exception as e:
                        db.session.rollback()

                        print("Error occurred:", str(e))
                        return {"error": str(e)}, 500
                    
                final_message = " & ".join(accumulated_message)
                return {"message": final_message}, 200
                                    
            elif data["typeof_action"] == "mapping_change":
                try:
                    # Extract details from the request body
                    action_val = data["action_val"][0]
                    id = data["id"]
                    process_area = action_val["process_area"]
                    source_system_cd = action_val["source_system_cd"]

                    #  Get the mapping_id from the Process_Area_Mapping table
                    mapping_result = db.session.execute(
                        select(self.process_area_mapping_table).where(
                            self.process_area_mapping_table.c.process_area
                            == process_area,
                            self.process_area_mapping_table.c.source_system_cd
                            == source_system_cd,
                            self.process_area_mapping_table.c.is_deleted == 0,
                        )
                    ).fetchone()

                    if not mapping_result:
                        return {"message": "Mapping not found."}, 404
                    mapping_dict = dict(mapping_result._mapping)
                    mapping_id = mapping_dict["mapping_id"]
                    print("Mapping ID found:", mapping_id)  # Debugging

                    #  Get org_id using org_name (highest level)
                    org_name = None
                    for level in range(1, 8):
                        level_key = f"level{level}"
                        if level_key in action_val and action_val[level_key]:
                            org_name = action_val[level_key]

                    if not org_name:
                        return {"message": "No organization name found in levels."},400

                    org_result = db.session.execute(
                        select(self.org_hierarchy_table).where(
                            self.org_hierarchy_table.c.org_name == org_name,
                            self.org_hierarchy_table.c.is_deleted == 0,
                            self.org_hierarchy_table.c.rec_end_date
                            == "9999-01-01 00:00:00.000",
                        )
                    ).fetchone()

                    if not org_result:
                        return {
                                    "message": "No active organization found with the specified name."
                                },404
                        

                    org_id = org_result.org_id
                    print("Organization ID found:", org_id)

                    # New record for Org_Hier_Mapping
                    new_mapping_record = {
                        "mapping_id": mapping_id,
                        "id": id,
                        "org_id": org_id,
                        "source_system_cd": source_system_cd,
                        "created_date": datetime.now(),
                        "created_by": current_user,
                        "updated_date": datetime.now(),
                        "updated_by": current_user,
                        "is_deleted": 0,
                        "rec_start_date": datetime.now(),
                        "rec_end_date": None,
                        "approval_1_by": "Pending",
                        "approval_1_date": None,
                        "approval_1_status": None,
                        "approval_2_by": None,
                        "approval_2_date": None,
                        "approval_2_status": "Pending",
                    }

                    # Step 5: Map dynamic fields based on incoming mapping keys
                    # Get the dynamic field names from the mapping result
                    dynamic_field_names = [
                        mapping_dict[f"dynamic_mapping_field_name_{i}"]
                        for i in range(1, 13)  # Assuming 12 dynamic fields
                    ]

                    print("Dynamic Field Names:", dynamic_field_names)

                    #  Mapping  the incoming keys to the dynamic fields
                    for key, value in action_val["mapping"].items():
                        if key in dynamic_field_names:
                            index = dynamic_field_names.index(key) + 1
                            dynamic_field_name = f"dynamic_mapping_field_name_{index}"
                            new_mapping_record[dynamic_field_name] = value
                            print(f"Mapping '{key}' to {dynamic_field_name}: {value}")

                    # db.session.execute(self.org_hier_mapping_table.insert().values(new_mapping_record))

                    # Extracting record id from org_mapping table to insert in workflow table
                    result = db.session.execute(
                        self.org_hier_mapping_table.insert()
                        .values(new_mapping_record)
                        .returning(self.org_hier_mapping_table.c.record_id)
                    )
                    record_id = result.scalar()
                    print("New Record ID:", record_id)

                    new_wid =  self.org_hier_mapping_helper.generate_workflow_id()

                    workflow_entry = db.session.execute(
                        select(self.workflow_static_table).where(
                            self.workflow_static_table.c.action == "update",
                            self.workflow_static_table.c.type == "single",
                            self.workflow_static_table.c.prop_type == "organization",
                            self.workflow_static_table.c.is_new == 0,
                        )
                    ).fetchone()

                    # Comparing the record nelwy created record with old record

                    request_record_id = data["record_id"]  # Extracted from request body
                    new_record_id = record_id  # The ID of the newly created record

                    # Step 1: Query the Org_Hier_Mapping table based on the two record IDs
                    query = select(self.org_hier_mapping_table).where(
                        (self.org_hier_mapping_table.c.record_id == request_record_id)
                        | (self.org_hier_mapping_table.c.record_id == new_record_id)
                    )
                    result = db.session.execute(query).fetchall()

                    changes = {}
                    if result:
                        records = [dict(row._mapping) for row in result]
                        old_record = next(
                            (r for r in records if r["record_id"] == request_record_id),
                            None,
                        )
                        new_record = next(
                            (r for r in records if r["record_id"] == new_record_id),
                            None,
                        )

                        if old_record and new_record:
                            # Compare org_id
                            if old_record["org_id"] != new_record["org_id"]:
                                old_org_id = old_record["org_id"]
                                new_org_id = new_record["org_id"]
                                print(
                                    f"Org ID changed from {old_org_id} to {new_org_id}"
                                )

                                # Query old and new org details
                                old_org_result = db.session.execute(
                                    select(self.org_hierarchy_table).where(
                                        self.org_hierarchy_table.c.org_id == old_org_id,
                                        self.org_hierarchy_table.c.is_deleted == 0,
                                    )
                                ).fetchone()
                                new_org_result = db.session.execute(
                                    select(self.org_hierarchy_table).where(
                                        self.org_hierarchy_table.c.org_id == new_org_id,
                                        self.org_hierarchy_table.c.is_deleted == 0,
                                    )
                                ).fetchone()

                                if old_org_result:
                                    old_org_dict = dict(old_org_result._mapping)
                                    for i in range(1, 8):  # Assuming there are 7 levels
                                        level_key = f"level{i}"
                                        changes[level_key] = {
                                            "old_value": old_org_dict.get(level_key),
                                            "new_value": None,  # New org values will be populated below
                                        }

                                if new_org_result:
                                    new_org_dict = dict(new_org_result._mapping)
                                    for i in range(1, 8):  # Assuming there are 7 levels
                                        level_key = f"level{i}"
                                        changes[level_key]["new_value"] = (
                                            new_org_dict.get(level_key)
                                        )

                            # Compare dynamic columns
                            dynamic_field_to_attr = {
                                f"dynamic_mapping_field_name_{i}": mapping_dict[
                                    f"dynamic_mapping_field_name_{i}"
                                ]
                                for i in range(1, 13)  # Assuming 12 dynamic fields
                            }

                            for (
                                dynamic_field_name,
                                attr_name,
                            ) in dynamic_field_to_attr.items():
                                old_value = old_record.get(dynamic_field_name)
                                new_value = new_record.get(dynamic_field_name)

                                if old_value != new_value:
                                    changes[attr_name] = {
                                        "old_value": old_value,
                                        "new_value": new_value,
                                    }

                        print("Changes detected:", changes)

                        for attr_name, change in changes.items():
                            work_flow_entries = {
                                "wid": new_wid,
                                "dyn_col_1": attr_name,
                                "dyn_col_2": change["old_value"],
                                "dyn_col_3": change["new_value"],
                                "created_date": db.func.now(),
                                "email_id" : get_user_from_token()[3],
                                "updated_date": None,
                                "is_deleted": 0,
                                "wf_status": "Pending",
                                "wf_reviewer_1_name": None,
                                "wf_reviewer_2_name": None,
                                "requested_by": current_user,
                                "wf_reviewer_1_status": "Pending",
                                "wf_reviewer_1_status_date": None,
                                "wf_reviewer_2_status": "Pending",
                                "wf_reviewer_2_status_date": None,
                                "typeof_action": "update",
                                "typeof_cr": "single",
                                "Comments": None,
                                "depedent_workflow_id": None,
                                "prop_id": None,
                                "org_hierarchy_record_id": None,
                                "mapping_record_id": new_record_id,
                                "prop_id": workflow_entry.prop_id,
                            }

                            db.session.execute(
                                self.workflow_table.insert().values(work_flow_entries)
                            )

                    db.session.commit()
                    return {
                        "message": f"Mapping for {process_area} has been successfully updated.",
                        "changes": changes,
                    }, 200

                except Exception as e:
                    db.session.rollback()
                    print("Error occurred:", str(e))  # Debugging error
                    return {"error": str(e)}, 500
        except Exception as e:
                return {"error": str(e)}, 500

    @user_or_admin_authentication_required
    def delete(self):
        try:
            org_mgt_data = request.get_json()
            org_hierarchy_record = db.session.execute(
                select(
                    self.org_hierarchy_table.c.org_level,
                    self.org_hierarchy_table.c.org_name,
                ).where(self.org_hierarchy_table.c.org_id == org_mgt_data["org_id"])
            ).fetchone()

            print("org_hierarchy_record....", org_hierarchy_record)

            if org_hierarchy_record:
                org_hierarchy_record = dict(org_hierarchy_record._mapping)
                # Dynamically get the column name
                column_name = f"level{org_hierarchy_record['org_level']}"
                column = getattr(self.org_hierarchy_table.c, column_name)

                # Construct the query
                query = select(self.org_hierarchy_table.c.org_name).where(
                    column == org_hierarchy_record["org_name"]
                )

                # Execute the query
                org_hierarchy_child_orgs = db.session.execute(query).fetchall()
                org_hierarchy_child_orgs = [
                    dict(each_child_org._mapping)["org_name"] 
                    for each_child_org in org_hierarchy_child_orgs if dict(each_child_org._mapping)["org_name"] != org_hierarchy_record['org_name']
                ]
                query = select(
                    self.org_hier_mapping_table.c.org_id,
                    self.org_hier_mapping_table.c.record_id,
                    self.org_hier_mapping_table.c.mapping_id,
                ).where(self.org_hier_mapping_table.c.org_id == org_mgt_data["org_id"])
                org_hierarchy_mapping_assoc = db.session.execute(query).fetchone()
                print('org_hierarchy_mapping_assoc....',dict(org_hierarchy_mapping_assoc._mapping))
                if org_hierarchy_child_orgs or org_hierarchy_mapping_assoc:
                    return {
                        "message": "There are child organization associated",
                        "associated_child_organizations": org_hierarchy_child_orgs,
                        "hierarchy_mapping_association":dict(org_hierarchy_mapping_assoc._mapping)
                    }, 400
                # Execute the update query
                db.session.execute(
                    self.org_hierarchy_table.update()
                    .where(self.org_hierarchy_table.c.org_id == org_mgt_data["org_id"])
                    .values(
                        is_deleted=1,
                    )
                )
                db.session.execute(
                    self.org_hier_mapping_table.update()
                    .where(
                        self.org_hierarchy_table.c.record_id
                        == org_hierarchy_record["record_id"]
                    )
                    .values(
                        is_deleted=1,
                    )
                )
                db.session.commit()
                return {
                    "message": f"{org_hierarchy_record['org_name']} has been successfully deleted"
                }, 200
            else:
                return {"message": "Invalid organization."}, 400
        except Exception as e:

            return {"error": str(e)}, 500


class OrgLevelManagement(Resource):

    def __init__(self):
        self.org_hierarchy = Base.metadata.tables["db_nxtgen.Org_Hierarchy"]
        self.organization_mapping_helper = OrgProcessMappingHelper()
        self.DEFAULT_REC_END_DATE = "9999-01-01 00:00:00.000"

    @user_or_admin_authentication_required
    def get(self):
        try:
            # Store request parameters
            include_pending = request.args.get("Include_Pending", "").lower()
            org_level = request.args.get("org_level")
            org_name = request.args.get("org_name")
            final_level_casecade = {"parents": [], "children": []}

            # Query for all records if org_level and org_name are not provided
            if include_pending and not org_level and not org_name:
                query = self.organization_mapping_helper.build_base_query(
                    include_pending
                )
                parent_org_records = db.session.execute(query).fetchall()
                list_parent_org_records = [
                    dict(record._mapping) for record in parent_org_records
                ]

                if list_parent_org_records:
                    return (
                        org_level_schemas.dump(
                            self.organization_mapping_helper._update_status_in_list_of_records(
                                list_parent_org_records
                            )
                        ),
                        200,
                    )
                else:
                    return {"message": "Invalid org_name, Record not found"}, 404

            # Process if org_level and org_name are provided
            else:
                org_level = int(org_level)
                # Query for the parent organization
                query = self.organization_mapping_helper.build_base_query(
                    include_pending
                ).where(self.org_hierarchy.c.org_name == org_name)
                parent_org_record = db.session.execute(query).fetchone()

                if not parent_org_record:
                    return {"message": "Invalid org_name, Record not found"}, 404

                # Extract parent org records
                parent_org_levels = [
                    value
                    for key, value in dict(parent_org_record._mapping).items()
                    if key.startswith("level") and value and value != org_name
                ]
                parents = []

                for each_parent_org in parent_org_levels:
                    query = self.organization_mapping_helper.build_base_query(
                        include_pending
                    ).where(self.org_hierarchy.c.org_name == each_parent_org)
                    parent_record = db.session.execute(query).fetchone()
                    if parent_record:
                        parents.append(dict(parent_record._mapping))

                # Query for child organizations
                org_name_pt = f"level{org_level - 1}"
                parent_org_column = getattr(self.org_hierarchy.c, org_name_pt)

                query = self.organization_mapping_helper.build_base_query(
                    include_pending
                ).where(
                    parent_org_column == org_name,
                    self.org_hierarchy.c.org_level == org_level,
                )
                child_org_records = db.session.execute(query).fetchall()
                child_org_records = [
                    dict(record._mapping) for record in child_org_records
                ]

                # Format and return the result
                if parents:
                    final_level_casecade["parents"] = org_level_schemas.dump(
                        self.organization_mapping_helper._update_status_in_list_of_records(
                            parents
                        )
                    )
                if child_org_records:
                    final_level_casecade["children"] = org_level_schemas.dump(
                        self.organization_mapping_helper._update_status_in_list_of_records(
                            child_org_records
                        )
                    )
                return final_level_casecade, 200

        except Exception as e:
            return {
                "message": "Something went wrong",
                "data": None,
                "error": str(e),
            }, 500

class ProcessKPI(Resource):
    #@user_or_admin_authentication_required
    def get(self):
        """
        get api for process area and kpi name
        :param:
        :return: list of process area and kpi name
        """
        try:
            vw_combined_kpi_master = db_schema +"."+"vw_combined_kpi_master"
            
            kpi_master = metadata.tables[vw_combined_kpi_master]
            process_item = filter_view_based_data(kpi_master,'view')
            if not process_item:
                return {"message": "KPI master data Not Found"}, 404
            process_kpi_name = []
            for data in process_item:
                params_dict = {}
                if data.process_area:
                    params_dict["process_area"] = data.process_area
                if data.kpi_name:
                    params_dict["kpi_name"] = data.kpi_name
                if data.kpi_name:
                    params_dict["kpi_id"] = data.kpi_id
                if data.kpi_source:
                    params_dict["kpi_source"] = data.kpi_source
                params_dict["is_non_roll_up"] = data.is_non_roll_up  
                params_dict["sector"] = data.sector if data.sector else None
                params_dict["aggregate_type"] = data.aggregate_type if data.aggregate_type else None
                params_dict["kpi_master_bulk_upload"] = data.kpi_master_bulk_upload if data.kpi_master_bulk_upload else None
                if params_dict not in process_kpi_name:
                    process_kpi_name.append(params_dict)
            
            return process_kpi_name, 200
        except Exception as e:
            # logging.exception("Exception in Admin role check ==>"+str(e))
            return {
                "message": "Something went wrong",
                "data": None,
                "error": str(e),
            }, 500


class ReportingKPI(Resource):
    @user_or_admin_authentication_required
    def get(self):
        """
        Get api for reporting period
        :param:
        :return: list of reporting period
        """
        kpi_input = Base.classes.Kpi_Input
        reporting_period_dict = {}
        reporting_period = []
        try:
            kpi_input_data = db.session.query(kpi_input).all()
            if not kpi_input_data:
                return {"message": "KPIs Not Found"}, 404
            for data in kpi_input_data:
                if data.reporting_period.strftime("%b-%Y") not in reporting_period:
                    reporting_period.append(data.reporting_period.strftime("%b-%Y"))
            reporting_period_dict["reporting_period"] = reporting_period
            print(reporting_period_dict)
            return reporting_period_dict, 200
        except Exception as e:
            return {"message": "Something Went Wrong", "error": str(e)}, 500


class KPISearchData(Resource):

    @user_or_admin_authentication_required
    def get(self):
        """
        Get api for fetching the table format data in these 3 tables Kpi_Master, Kpi_Input, Org_Hierarchy
        :param : KpiId, Status, OrgId, ProcessArea, ReportingPeriod
        :return: reurning table format data
        """
        # Organization table
        org_hierarchy = Base.metadata.tables['db_nxtgen.Org_Hierarchy']
        # Kpi_master table
        kpi_master = Base.classes.Kpi_Master
        kpi_comment = Base.classes.Kpi_Comments_manualautomatic
        # Kpi_Input table
        kpi_input = Base.classes.Kpi_Input
        org_id = request.args.get("OrgId")
        process_area = request.args.get('ProcessArea')
        kpi_id = request.args.get('KpiId')
        sector = request.args.get('Sector')
        aggregate_type = request.args.get('aggregate_type')
        reporting_period = request.args.get('ReportingPeriod')
        fetch_sector_role_instance = FetchSectorRole()
        role_matrix, status_code = fetch_sector_role_instance.get()
        if isinstance(role_matrix, dict):
            sectors_with_roles = [entry["sector"] for entry in role_matrix["roleMatrix"] if entry["role"] == "Reviewer"]
        else:
            sectors_with_roles = []

        status = request.args.get('Status')
        if status:
            status = status.capitalize()
        org_ids = []
        process_area_list = []
        kpi_ids = []
        reporting_list = []
        if process_area:
            process_area_list.extend(process_area.split(','))
        
        if org_id:
            org_ids.extend(org_id.split(','))
        
        if kpi_id:
            kpi_ids.extend(kpi_id.split(','))
        
        if reporting_period:
            list_reporting = reporting_period.split(',')
            for reporing in list_reporting:
                reporing_date = datetime.strptime(reporing, '%b-%Y').date()
                reporing_date = reporing_date.strftime(dateformatStr)
                reporting_list.append(datetime.strptime(reporing_date, dateformatStr).date())

        print("reporting_list", reporting_list)
        # Create aliases for the models
        org_alias = aliased(org_hierarchy)
        kpi_alias = aliased(kpi_master)
        kpi_input_alias = aliased(kpi_input)
        comments_alias = aliased(kpi_comment)
        # Query for fetching data from 3 tables
        query = db.session.query(
            kpi_input_alias.id,
            org_alias.c.org_name,
            org_alias.c.org_id,
            org_alias.c.rec_end_date,
            kpi_alias.kpi_name,
            kpi_alias.kpi_id,
            kpi_alias.process_area,
            kpi_alias.kpi_type,
            kpi_alias.sector,
            kpi_alias.aggregate_type,
            kpi_alias.red_threshold,
            kpi_alias.green_threshold,
            kpi_alias.threshold_type,
            kpi_input_alias.reporting_period,
            kpi_input_alias.numerator,
            kpi_input_alias.denominator,
            kpi_input_alias.updated_date,
            kpi_input_alias.updated_by,
            kpi_input_alias.status,
            comments_alias.comment_id,
            kpi_input_alias.period_expired,
            kpi_input_alias.value,
            kpi_input_alias.kpi_Input_bulk_upload). \
            join(kpi_alias, kpi_alias.kpi_id == kpi_input_alias.kpi_id). \
            join(org_alias, org_alias.c.org_id == kpi_input_alias.org_id). \
            outerjoin(comments_alias, and_(comments_alias.org_id == kpi_input_alias.org_id,comments_alias.kpi_id == kpi_input_alias.kpi_id,comments_alias.reporting_period == kpi_input_alias.reporting_period)).filter(org_alias.c.rec_end_date == '9999-01-01 00:00:00.000')


        # Apply filters if criteria are not empty
        if org_ids:
            query = query.filter(kpi_input_alias.org_id.in_(org_ids))

        if process_area_list:
            query = query.filter(kpi_alias.process_area.in_(process_area_list))

        if kpi_ids:
            query = query.filter(kpi_alias.kpi_id.in_(kpi_ids))
        if reporting_list:
            query = query.filter(kpi_input_alias.reporting_period.in_(reporting_list))
        if status:
            if status != 'Pending':
                query = query.filter(kpi_input_alias.status == status)
            else :
                query = query.filter(kpi_input_alias.status == status).filter(kpi_alias.sector.in_(sectors_with_roles))
        if sector:
            query = query.filter(kpi_alias.sector == sector)
        if aggregate_type:
            query = query.filter(kpi_alias.aggregate_type == aggregate_type)
        kpi_data = query.all()
        
        kpi_data = search_schemas.dump(kpi_data)
        list_kpi_data = []
        for data in kpi_data:
            reporting_date = datetime.strptime(data["reporting_period"], dateformatStr)
            data["reporting_period"] = reporting_date.strftime('%b-%Y')
            list_kpi_data.append(data)
        original_kpi_data = max_day_and_days_difference(list_kpi_data, kpi_input)
        return original_kpi_data, 200
        

# update api for updating data
class KPIInputUpdate(Resource):
    @user_or_admin_authentication_required
    def put(self):
        """
        Put api for updating the numerator and denominator in Kpi Input table.
        :Body: json format- id,numerator, denominator, status
        :return: Updated successfully
        """
        try:
            kpi_input = Base.classes.Kpi_Input
            data = request.get_json()
            kpi_id = data["id"]  # Renamed variable to avoid shadowing built-in 'id'
            kpi_data = db.session.query(kpi_input).get(kpi_id)
            if not kpi_data:
                return {"message": "KPI data not found with ID:" + str(kpi_id)}, 404
            is_user_found, current_user, user_role = get_user_from_token()
            if is_user_found:
                print("current user", current_user)
            if "numerator" in data and "denominator" in data:
                kpi_data.numerator = data["numerator"]
                kpi_data.denominator = data["denominator"]
                kpi_data.value = data["value"]
                kpi_data.updated_by = current_user
                kpi_data.updated_date = db.func.now()

                kpi_data.status = data["status"].capitalize()

                db.session.commit()
                updated_kpi_input_row = (
                    db.session.query(kpi_input).filter_by(id=kpi_id).first()
                )
                kpi_input_history_data = kpi_input_schema.dump(updated_kpi_input_row)
                insert_kpi_input_history(kpi_input_history_data)

                return kpi_input_schema.dump(kpi_data), 200
            else:
                return {"message": "Data not found"}, 400
        except Exception as e:
            return {"message": "Something Went wrong", "error": str(e)}, 500


class KPIInputHistory(Resource):

    def __init__(self):
        super().__init__()

    @user_or_admin_authentication_required
    def get(self):
        """
        Get api to fetch kpi_input history
        :param: kpi_id,org_id,reporting_period
        :return: list of kpi_input history
        """
        try:
            request_args = {
                "org_id": request.args.get("org_id"),
                "kpi_id": request.args.get("kpi_id"),
                "reporting_period": request.args.get("reporting_period"),
            }
            if any(value is None for value in request_args.values()):
                return {
                    "message": f"{next((key for key, value in request_args.items() if value is None), None)} is a mandatory field",
                }, 400
            else:
                fetch_kpi_input_histroy_data = fetch_kpi_input_histroy(request_args)
                return fetch_kpi_input_histroy_data, 200
        except Exception as e:
            return {"message": "Something went wrong", "error": str(e)}, 500


class KPIApprovalResource(Resource):
    @admin_authentication_required
    def post(self):
        kpi_input = Base.classes.Kpi_Input
        data = request.get_json()

        kpi_input_id_list = data["kpi_input_id_list"]

        current_user = get_user_from_token()[1]

        for kpi_input_id in kpi_input_id_list:
            try:
                kpi_input_record = (
                    db.session.query(kpi_input).filter_by(id=kpi_input_id).first()
                )
            except Exception as e:
                return {
                    "message": "KPI Input Not Found with ID:" + str(kpi_input_id)
                }, 404

            if not kpi_input_record:
                return {
                    "message": "KPI Input Not Found with ID:" + str(kpi_input_id)
                }, 404
            else:
                kpi_input_record.status = "Approved"
                kpi_input_record.updated_by = current_user
                kpi_input_record.updated_date = db.func.now()

        db.session.commit()

        return {"message": "All KPI Inputs Approved"}, 200


BLANK_FIELD_ERROR_MSG = "This field cannot be left blank!"
kpi_input_create_args = reqparse.RequestParser()
kpi_input_create_args.add_argument(
    "kpi_id", type=int, required=True, help=BLANK_FIELD_ERROR_MSG
)
kpi_input_create_args.add_argument(
    "org_id", type=int, required=True, help=BLANK_FIELD_ERROR_MSG
)
kpi_input_create_args.add_argument(
    "reporting_period", type=str, required=True, help=BLANK_FIELD_ERROR_MSG
)
kpi_input_create_args.add_argument(
    "numerator", type=str, required=True, help=BLANK_FIELD_ERROR_MSG
)
kpi_input_create_args.add_argument(
    "denominator", type=str, required=True, help=BLANK_FIELD_ERROR_MSG
)
kpi_input_create_args.add_argument(
    "value", type=str, required=True, help=BLANK_FIELD_ERROR_MSG
)
kpi_input_create_args.add_argument(
    "status", type=str, required=True, help=BLANK_FIELD_ERROR_MSG
)


class KPIInputCreate(Resource):
    @user_or_admin_authentication_required
    def post(self):
        """
        POST api for creating new Kpi Input record.
        :Body: JSON format- kpi_id,org_id,reporting_period,numerator,denominator,value,status
        :returns: JSON format- of KPI Input record.
        """

        data = kpi_input_create_args.parse_args()
        try:
            kpi_input = Base.classes.Kpi_Input
            current_user = get_user_from_token()[1]

            reporting_period = data["reporting_period"]
            reporting_period_date_obj = datetime.strptime(
                reporting_period, "%b-%Y"
            ).date()

            data["reporting_period"] = reporting_period_date_obj
            data["status"] = data["status"].capitalize()
            data["created_by"] = current_user
            data["updated_by"] = current_user
            data["created_date"] = db.func.now()
            data["updated_date"] = db.func.now()
            data["is_deleted"] = False
            data["period_expired"] = False

            new_kpi_input = kpi_input(**data)

            db.session.add(new_kpi_input)
            db.session.commit()
            inserted_kpi_input_row = (
                db.session.query(kpi_input)
                .filter_by(id=kpi_input_schema.dump(new_kpi_input)["id"])
                .first()
            )
            kpi_input_history_data = kpi_input_schema.dump(inserted_kpi_input_row)
            insert_kpi_input_history(kpi_input_history_data)

            return kpi_input_schema.dump(new_kpi_input), 201
        except IntegrityError as e:
            return {
                "message": f"Exception, duplicate combination found - kpi_id - {data['kpi_id']}, org_id - {data['org_id']}, reporting_period - {data['reporting_period']}"
            }, 500
        except Exception as e:
            return {"message": "Exception, while creating KPI Input -" + str(e)}, 500


class KPIInputBulkUpdate(Resource):
    @user_or_admin_authentication_required
    def post(self):
        """
        POST api for bulk update or insert of KPI Input records.
        :Body: JSON array of records, each having attributes id, kpi_id, org_id, reporting_period, numerator, denominator, value, status
        :returns: JSON format- status of operation for each record.
        """
        data = request.get_json()

        if not isinstance(data, list):
            return {"message": "Invalid input format, expected a JSON array"}, 400
        for record in data:
            result = self.process_record(record)
            if result["status"] == "Record not found, update failed":
                return {
                    "message": "Record not found, update failed with id:-"
                    + str(result["id"])
                }, 400

        return {"message": "KPI Input Added/Updated Successfully"}, 200

    def process_record(self, record):
        """
        Process each record in the JSON array.
        If the record contains an 'id' attribute, it updates the existing record.
        Otherwise, it inserts a new record.
        """
        current_user = get_user_from_token()[1]
        if 'id' in record:
            id = record["id"]
            kpi_input = Base.classes.Kpi_Input
            kpi_data = db.session.query(kpi_input).get(id)
            
            if not kpi_data:
                return {"id": id, "status": "Record not found, update failed"}

            if "numerator" in record and "denominator" in record:
                kpi_data.numerator = record["numerator"]
                kpi_data.denominator = record["denominator"]

            if "reporting_period" in record:
                reporting_period_date_obj = datetime.strptime(
                    record["reporting_period"], "%b-%Y"
                ).date()
                kpi_data.reporting_period = reporting_period_date_obj

            if "value" in record:
                kpi_data.value = record["value"]

            if "status" in record:
                kpi_data.status = record["status"].capitalize()

            kpi_data.updated_by = current_user
            kpi_data.updated_date = db.func.now()
            db.session.commit()
            updated_kpi_input_row = db.session.query(kpi_input).filter_by(id=id).first()
           
            kpi_input_history_data=kpi_input_schema.dump(updated_kpi_input_row)
            insert_kpi_input_history(kpi_input_history_data)
            return {"id": id, "status": "Record updated successfully"}
        else:
            if "reporting_period" in record:
                reporting_period_date_obj = datetime.strptime(
                    record["reporting_period"], "%b-%Y"
                ).date()
                record["reporting_period"] = reporting_period_date_obj
                record["status"] = record["status"].capitalize()
                record["created_by"] = current_user
                record["updated_by"] = current_user
                record["created_date"] = db.func.now()
                record["updated_date"] = db.func.now()
                record["is_deleted"] = False
                record["period_expired"] = False
            kpi_input = Base.classes.Kpi_Input

            new_kpi_input = kpi_input(**record)
            db.session.add(new_kpi_input)
            db.session.commit()
            inserted_kpi_input_row = (
                db.session.query(kpi_input)
                .filter_by(id=kpi_input_schema.dump(new_kpi_input)["id"])
                .first()
            )
            kpi_input_history_data = kpi_input_schema.dump(inserted_kpi_input_row)

            insert_kpi_input_history(kpi_input_history_data)
            return {"status": "New record inserted successfully"}


class KPIBookmarkFeature(Resource):

    def __init__(self):

        self.book_mark_obj = kpi_manual_input()
    @user_or_admin_authentication_required
    def post(self):
        """
        POST api to save bookmark in db.
        :Body: JSON array of records, each having attributes bookmark_type_id, bookmark_name, oid, attribute,
        :returns: JSON format- success message of insertion.
        """
        return self.book_mark_obj.save_book_mark(
            request.get_json(), request.headers["Authorization"].split(" ")[1]
        )

    @user_or_admin_authentication_required
    def delete(self):
        """
        GET api to fetch bookmark records.
        :param: query param -oid,
        :returns: JSON format- array of bookmark records.
        """
        return self.book_mark_obj.delete_book_mark(request.args.get("bookmark_id"))

    @user_or_admin_authentication_required
    def put(self):
        """
        GET api to fetch bookmark records.
        :param: query param -oid,
        :returns: JSON format- array of bookmark records.
        """

        return self.book_mark_obj.update_book_mark(request.get_json())

    @user_or_admin_authentication_required
    def get(self):
        """
        GET api to fetch detailed bookmark records.
        :param: query param -bookmark_id ,
        :returns: JSON format- array of bookmark records.
        """
        return self.book_mark_obj.fetch_book_mark_view(request.args.get("bookmark_id"))

    @user_or_admin_authentication_required
    def patch(self):
        """
        GET api to fetch detailed bookmark records.
        :param: query param -bookmark_id ,
        :returns: JSON format- array of bookmark records.
        """
        return self.book_mark_obj.rename_bookmark(request.get_json())


class KPIBookmarkFeatureGetAll(Resource):

    def __init__(self):

        self.book_mark_obj = kpi_manual_input()
    
    @user_or_admin_authentication_required
    def get(self):
        """
        GET api to fetch bookmark records.
        :param: query param -oid,
        :returns: JSON format- array of bookmark records.
        """
        if len(request.args) == 0:
            return {"type": "oid", "message": "This field cannot be left blank!"}
        else:
            return self.book_mark_obj.fetch_book_mark(request.args)


class BulkUploadExcelParser(Resource):

    def __init__(self):
        super().__init__()
        self.kpi_input = Base.classes.Kpi_Input
        self.is_user_found, self.current_user, self.user_role = get_user_from_token()

    @admin_authentication_required
    def post(self):
        if 'file' not in request.files:
            return "No file part", 400
        file = request.files['file']
        if file.filename == '':
            return "No selected file", 400
        sectors = request.form.getlist("sector")
        if not sectors:
            return 'No sector provided', 400
        sectors_string = ','.join(sectors)
        sectors_string = sectors_string.replace('"', '')
        allowed_sectors = sectors_string.split(',')
        response_buffer, errored_rows_number, inserted_rows_number = self.process_file(file, allowed_sectors)
        
        if isinstance(response_buffer, dict):
            return response_buffer, 400
        
        if errored_rows_number > 0:
            response = make_response(send_file(response_buffer, as_attachment=True, download_name='validated_output.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
            response.headers['Errored-Rows-Number'] = str(errored_rows_number)
            response.headers['Inserted-Rows-Number'] = str(inserted_rows_number)
            response.headers['Access-Control-Expose-Headers'] = 'Errored-Rows-Number, Inserted-Rows-Number'
            return response
        else:
            response = make_response()
            response.headers['Errored-Rows-Number'] = str(errored_rows_number)
            response.headers['Inserted-Rows-Number'] = str(inserted_rows_number)
            response.headers['Access-Control-Expose-Headers'] = 'Errored-Rows-Number, Inserted-Rows-Number'
            return response
 
    def validate_data(self, df, allowed_sectors):
        validation_errors = []

        for idx, row in df.iterrows():
            error = []
            for field in [
                "sector",
                "process_area",
                "kpi_type",
                "kpi_name",
                "org_name(Exact Name of any Level 1 - 6)",
                "reporting_period",
            ]:
                if pd.isnull(row[field]):
                    error.append(f"{field} cannot be null")
            if row["sector"] not in allowed_sectors:
                error.append("You do not have access to this KPI.")

            if row["kpi_type"] == "percentage":
                if pd.isnull(row["numerator"]) or pd.isnull(row["denominator"]):
                    error.append(
                        "Numerator or Denominator field cannot be null  for percentage KPI"
                    )
                else:
                    if row["denominator"] == 0 and row["numerator"] > 0:
                        error.append(
                            "Denominator cannot be 0 if Numerator is greater than 0 for percentage KPI"
                        )
            elif row["kpi_type"] == "absolute number":
                if not pd.isnull(row["numerator"]) or not pd.isnull(row["denominator"]):
                    error.append(
                        "Numerator and Denominator field should be null for absolute KPI"
                    )
                if pd.isnull(row["value"]):
                    error.append("Value  field cannot be null for absolute number KPI")

            validation_errors.append("; ".join(error) if error else None)

        df["validation_errors"] = validation_errors
        return df

    def highlight_invalid_rows(self, workbook, df):
        ws = workbook.active

        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        for idx, row in df.iterrows():
            if row["validation_errors"]:
                for col in range(1, len(row) + 1):
                    ws.cell(row=idx + 2, column=col).fill = red_fill
                ws.cell(row=idx + 2, column=len(row) + 1).value = row[
                    "validation_errors"
                ]

    def query_database(self):
        vw_combined_kpi_master = db_schema + "." + "vw_combined_kpi_master"
        kpi_master = metadata.tables[vw_combined_kpi_master]
        kpi_query = db.session.query(
            kpi_master.columns.kpi_name,
            kpi_master.columns.sector,
            kpi_master.columns.kpi_id,
            kpi_master.columns.kpi_type,
            kpi_master.columns.is_non_roll_up,
            kpi_master.columns.red_threshold,
            kpi_master.columns.green_threshold,
        ).distinct()
        kpi_id_rows = pd.DataFrame(
            kpi_query.all(),
            columns=[
                "kpi_name",
                "sector",
                "kpi_id",
                "kpi_type",
                "is_non_roll_up",
                "red_threshold",
                "green_threshold",
            ],
        )

        org_hierarchy = Base.metadata.tables["db_nxtgen.Org_Hierarchy"]
        org_query = db.session.query(
            org_hierarchy.c.org_name, org_hierarchy.c.org_id
        ).distinct()
        org_id_rows = pd.DataFrame(org_query.all(), columns=["org_name", "org_id"])

        return kpi_id_rows, org_id_rows

    def match_kpi_org(self, df, kpi_df, org_df, is_non_roll_up):
        for idx, row in df.iterrows():
            if not row["validation_errors"]:
                if (
                    "kpi_name" not in kpi_df.columns
                    or "kpi_type" not in kpi_df.columns
                    or "sector" not in kpi_df.columns
                ):
                    continue
                kpi_match = kpi_df[
                    (kpi_df["kpi_name"] == row["kpi_name"])
                    & (kpi_df["sector"] == row["sector"])
                    & (kpi_df["is_non_roll_up"].isnull())
                ]
                kpi_match_non_roll_up = kpi_df[
                    (kpi_df["kpi_name"] == row["kpi_name"])
                    & (kpi_df["sector"] == row["sector"])
                ]

                org_match = org_df[
                    (
                        org_df["org_name"]
                        == row["org_name(Exact Name of any Level 1 - 6)"]
                    )
                ]
                if kpi_match.empty and kpi_match_non_roll_up.empty:
                    row["validation_errors"] = (
                        row["validation_errors"] or ""
                    ) + "KPI not found; "
                elif (
                    kpi_match.empty
                    and not is_non_roll_up
                    and not kpi_match_non_roll_up.empty
                ):
                    row["validation_errors"] = (
                        row["validation_errors"] or ""
                    ) + "You do Not have access to Non-Roll Up KPI"
                if org_match.empty:
                    row["validation_errors"] = (
                        row["validation_errors"] or ""
                    ) + "Organization not found; "

                if row["validation_errors"]:
                    df.at[idx, "validation_errors"] = row["validation_errors"]

        return df

    def generate_objects(self, df, kpi_id_rows, org_id_rows):
        objects = []
        error_indices = []

        for idx, row in df.iterrows():
            if pd.isnull(row["validation_errors"]) or not row["validation_errors"]:
                kpi_row = kpi_id_rows[
                    (kpi_id_rows["kpi_name"] == row["kpi_name"])
                    & (kpi_id_rows["sector"] == row["sector"])
                ]

                if not kpi_row.empty:
                    kpi_type = row.get("kpi_type")
                    numerator = row.get("numerator")
                    denominator = row.get("denominator")

                    if kpi_type == "percentage":
                        if numerator == 0 and denominator == 0:
                            value = None
                        elif (
                            numerator is not None
                            and denominator is not None
                            and denominator != 0
                        ):
                            value = (numerator / denominator) * 100
                        else:
                            value = None
                    else:
                        value = row.get("value", None)

                    org_id = org_id_rows[
                        (
                            org_id_rows["org_name"]
                            == row["org_name(Exact Name of any Level 1 - 6)"]
                        )
                    ]
                    reporting_period = row["reporting_period"]

                    if isinstance(reporting_period, str):
                        reporting_period = datetime.strptime(
                            reporting_period, "%m/%d/%Y"
                        )
                        formatted_reporting_period = reporting_period.strftime(
                            dateformatStr
                        )
                    elif isinstance(reporting_period, datetime):
                        formatted_reporting_period = reporting_period

                    current_date = datetime.now()
                    three_months_later = reporting_period + timedelta(days=3 * 30)
                    if current_date > three_months_later:
                        period_expired = 1
                    else:
                        period_expired = 0
                    created_date = datetime.now()
                    updated_date = datetime.now()

                    if row.get("numerator") is None or math.isnan(row.get("numerator")):
                        numerator = None
                    else:
                        numerator = float(row["numerator"])

                    if row.get("denominator") is None or math.isnan(
                        row.get("denominator")
                    ):
                        denominator = None
                    else:
                        denominator = float(row["denominator"])

                    obj = {
                        "kpi_id": int(kpi_row.iloc[0]["kpi_id"]),
                        "org_id": int(org_id.iloc[0]["org_id"]),
                        "reporting_period": formatted_reporting_period,
                        "numerator": numerator,
                        "denominator": denominator,
                        "created_date": created_date,
                        "created_by": self.current_user,
                        "updated_date": updated_date,
                        "updated_by": self.current_user,
                        "status": "Approved",
                        "is_deleted": 0,
                        "value": float(value) if value is not None else None,
                        "period_expired": period_expired,
                        "kpi_Input_bulk_upload": "Y",
                    }

                    try:
                        kpi_input = Base.classes.Kpi_Input
                        new_kpi_input = self.kpi_input(**obj)
                        db.session.add(new_kpi_input)
                        db.session.commit()
                        inserted_kpi_input_row = (
                            db.session.query(kpi_input)
                            .filter_by(id=kpi_input_schema.dump(new_kpi_input)["id"])
                            .first()
                        )
                        kpi_input_history_data = kpi_input_schema.dump(
                            inserted_kpi_input_row
                        )
                        insert_kpi_input_history(kpi_input_history_data)
                    except IntegrityError as e:
                        db.session.rollback()
                        error_indices.append(idx)

        for idx in error_indices:
            df.loc[idx, "validation_errors"] = "Data already exists"
        objects = [obj for idx, obj in enumerate(objects) if idx not in error_indices]
        return objects

    def get_nonrollup_authority(self):
        self.nonruollupAuthorityObj = KPINonRollup()
        response_data = self.nonruollupAuthorityObj.get()
        if isinstance(response_data, dict) and "message" in response_data:
            return False
        else:
            return True

    # Index Method for the File Processing.
    # validate_data      - Validates each row and adds appropriate error if found invalid.
    # query_database     - Queries the vw_combined_kpi_master view to fetch all KPIs and org_hierarchy to get all Organizations
    # match_kpi_org      - Loops through and finds the matching kpi row with the kpi_name and sector combination and extracts the kpi_id.
    #                     It also validates that if no kpi_id is found it adds error as kpi not found. Similar operation is done for
    #                     org_herirarchy to extract and validate org_id.
    # generate_objects   - This method is responsible for calculating the Value, type casting the necessary column fields to match
    #                     the DB columns data types. It also formats the reporting period appropriately and finally creates the
    #                     object to be inserted in the kpi_input table. It commits each object and handle intergrity error so that
    #                     duplicate rows can also be added to the error row list. It also inserts data in the kpi_input_history_table
    #                     for successfully parsed data.
    def process_file(self, file, allowed_sectors):
        df = pd.read_excel(
            file, sheet_name=1 if len(pd.ExcelFile(file).sheet_names) > 1 else 0
        )
        df["kpi_type"] = df["kpi_type"].str.lower()
        df = self.validate_data(df, allowed_sectors)
        is_non_roll_up = self.get_nonrollup_authority()
        kpi_id_rows, org_id_rows = self.query_database()
        df = self.match_kpi_org(df, kpi_id_rows, org_id_rows, is_non_roll_up)
        self.generate_objects(df, kpi_id_rows, org_id_rows)
        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        workbook = load_workbook(output)
        self.highlight_invalid_rows(workbook, df)
        response_buffer = BytesIO()
        workbook.save(response_buffer)
        response_buffer.seek(0)
        return (
            response_buffer,
            df["validation_errors"].notnull().sum(),
            df["validation_errors"].isnull().sum(),
        )


class GetHierarchyMappingSourceData(Resource):

    def __init__(self):
        # Initialize the table reference in the constructor
        self.process_area_mapping = metadata.tables["db_nxtgen.Process_Area_Mapping"]

    @user_or_admin_authentication_required
    def get(self):
        try:
            source_system_cd = request.args.get("source_system_cd")

            if not source_system_cd:
                # Retrieve distinct source_system_cd and source_system_name
                query = (
                    select(
                        self.process_area_mapping.c.source_system_cd,
                        self.process_area_mapping.c.source_system_name
                    )
                    .distinct()
                    .where(self.process_area_mapping.c.is_deleted == 0)
                )

                qry_result = db.session.execute(query).fetchall()
                
                # Build a list of dicts that include both the code and name
                data = [
                    {
                        "source_system_cd": row._mapping["source_system_cd"],
                        "source_system_name": row._mapping["source_system_name"],
                    }
                    for row in qry_result
                ]

                return {
                    "message": "Source data retrieved successfully",
                    "data": data,
                }, 200

            else:
                # Retrieve process areas for the selected source_system_cd
                query = (
                    select(
                        self.process_area_mapping.c.process_area,
                        self.process_area_mapping.c.source_system_name,
                    )
                    .where(
                        self.process_area_mapping.c.source_system_cd == source_system_cd,
                        self.process_area_mapping.c.is_deleted == 0,
                    )
                )

                qry_result = db.session.execute(query).fetchall()

                if qry_result:
                    # Process areas and source system name
                    process_areas = [{"process_area": row._mapping["process_area"]} for row in qry_result]

                    return {
                        "message": "Process Area retrieved successfully",
                        "data": process_areas,
                    }, 200
                else:
                    return {
                        "message": "No process area found for the selected source system",
                        "data": [],
                    }, 404
        except Exception as e:
            return {
                "message": "Something went wrong",
                "data": None,
                "error": str(e),
            }, 500


class GetHierarchyMappingProcessArea(Resource):

    def __init__(self):
        # Initialize the table reference in the constructor
        self.process_area_mapping = Base.metadata.tables[
            "db_nxtgen.Process_Area_Mapping"
        ]

    @user_or_admin_authentication_required
    def get(self):
        """
        GET API to fetch distinct process areas from the Process_Area_Mapping table.
        :return: A list of process areas for the dropdown.
        """
        try:
            query = (
                select(
                    self.process_area_mapping.c.process_area,
                    self.process_area_mapping.c.source_system_cd,
                    self.process_area_mapping.c.source_system_name,
                )
                .distinct()
                .where(self.process_area_mapping.c.is_deleted == 0)
            )

            result = db.session.execute(query).fetchall()
            if result:
                # Here extracting the process_area values and forms a list
                process_area_list = [dict(row._mapping) for row in result]

                return {"process_area": process_area_list}, 200
            else:
                return {"message": "No process area found", "process_area": []}, 404

        except Exception as e:
            return {
                "message": "Something went wrong",
                "data": None,
                "error": str(e),
            }, 500
